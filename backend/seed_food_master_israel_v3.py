"""Full-REPLACE seed script for the curated Israeli food_master catalog
(2026-08-01), built from a multi-sheet .xlsx (not the flat CSV pair
seed_food_master_v2.py consumed).

Usage (run from backend/ directory, with the correct DATABASE_URL for the
target environment already set):

    python seed_food_master_israel_v3.py <path_to_xlsx> [--dry-run]

Same PRESERVED_HE_NAMES carve-out as seed_food_master_v2.py (imported from
there directly, so both scripts always agree on the exact set of legacy
rows this applies to): those 9 legacy (fdc_id IS NULL) rows are left
active and untouched. Every other existing food_master row whose fdc_id
is not in this file's "מזונות ישראל" sheet gets is_active=False (never
hard-deleted -- FoodLog is free-text with no FK to food_master, so
nothing breaks, but the historical row stays queryable). Final active
catalog: 334 new rows + 9 preserved legacy rows = 343 (decision confirmed
2026-08-01 -- an earlier version of this script did a literal full
replace with no carve-out; that deactivated the 9 legacy rows too, since
none of them have an equivalent anywhere in the new file even by name).

Sheets consumed:
  מזונות ישראל       -- the 334-row catalog itself (338 rows on the sheet,
                         last 4 blank padding -- filtered on name is None)
  מנות ישראל         -- food_portions rows, keyed by fdc_id like v2
  מיפוי שמות         -- per-fdc_id alias ("שם קצר לחיפוש"), skipped for any
                         row whose פעולה starts with "הוסר" (both
                         "הוסר ככפילות" and "הוסר — לא מתאים" appear)
  יומן שינויים        -- informational only; removals are already baked
                         into "מזונות ישראל" (not present there = correctly
                         excluded), nothing to act on here.

macros_estimated is hardcoded False for every row this script writes.
"דוח השלמת מאקרו" was originally going to drive this (rows whose "פירוט"
mentions Atwater), but that conflates USDA's own standard costing
methodology -- Atwater factors are how USDA computes energy for a huge
share of its own database, not something specific or suspect about these
particular rows -- with FitAI actually estimating a value itself. Under
the narrower, correct definition ("FitAI calculated this with no USDA
figure backing it"), nothing in this file qualifies: every value is
either original, a direct FDC field recovery, or matched to an
equivalent USDA SR Legacy/FNDDS record -- all USDA-sourced either way
(decision confirmed 2026-08-01; the ~198-row Atwater-detail figure from
the first dry run measured the wrong thing).
"""
import sys

import openpyxl

from app.core.database import SessionLocal
from app.models.fitness import FoodMaster, FoodPortion
from seed_food_master_v2 import PRESERVED_HE_NAMES

SHEET_CATALOG = "מזונות ישראל"
SHEET_PORTIONS = "מנות ישראל"
SHEET_ALIASES = "מיפוי שמות"

# 17 category strings actually present in this file's "קטגוריה" column ->
# the app's existing 7-value FoodCategory enum (schemas/food.py), same
# taxonomy and same category_en-preserves-the-original approach as
# seed_food_master_v2.py's CATEGORY_MAP. Several source labels differ
# slightly from the USDA-overhaul file's wording (e.g. "דגים" here vs
# "דגים ופירות ים" there, "נקניקים" here vs "נקניקים ובשרי מעדנייה"
# there) -- mapped to the same target either way.
CATEGORY_MAP = {
    "ירקות ומוצרי ירקות": "ירקות",
    "פירות ומיצי פירות": "פירות",
    "מוצרי חלב וביצים": "מוצרי חלב",
    "משקאות": "מוצרי חלב",
    "דגנים ופסטה": "פחמימות",
    "מוצרי מאפה": "פחמימות",
    "ממתקים": "פחמימות",
    "קטניות ומוצרי קטניות": "חלבונים",
    "דגים": "חלבונים",
    "מוצרי בקר": "חלבונים",
    "נקניקים": "חלבונים",
    "מוצרי עוף": "חלבונים",
    "כבש, עגל ובשר ציד": "חלבונים",
    "שומנים ושמנים": "שומנים",
    "אגוזים וזרעים": "שומנים",
    "תבלינים ועשבי תיבול": "שומנים",
    "מרקים, רטבים ורוטבי צלי": "ירקות",
}

# The portions sheet only has one (Hebrew) unit column -- food_portions.unit_en
# is NOT NULL, so every distinct "יחידה" value actually seen in the file
# needs an English counterpart. Falls back to reusing the Hebrew string
# itself (still satisfies the column, just not actually English) for
# anything not in this table -- seed() reports any fallback hits so they
# can be added here later instead of silently shipping non-English text.
UNIT_HE_TO_EN = {
    "כוס": "cup", "יחידה אחת": "1 unit", "יחידה": "unit", "כף": "tbsp",
    "כפית": "tsp", "אונקיה": "oz", "פרי": "fruit", "מיליליטר": "ml",
    "מנה": "serving", "מנה מבושלת": "cooked serving", "פרוסה": "slice",
    "נתח צלי": "roast cut", "סטייק": "steak", "פלח": "wedge", "ביצה": "egg",
    "חתיכה": "piece", "חתיכות": "pieces", "מכל": "container", "פילה": "fillet",
    "קופסת שימורים": "can", "נקניקייה": "sausage", "בצל": "onion",
    "בננה": "banana", "עגבניות": "tomato", "זית": "olive", "עוגייה": "cookie",
    "צרור": "bunch", "אריזה": "package",
}


def _cell_str(ws, row, col):
    value = ws.cell(row, col).value
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def _cell_float(ws, row, col):
    value = ws.cell(row, col).value
    return float(value) if value is not None else None


def _load_aliases(ws):
    """fdc_id -> alias string, skipping any row whose פעולה starts with
    'הוסר' ('הוסר ככפילות' / 'הוסר — לא מתאים' both seen in the file --
    matched by prefix so any future 'הוסר ...' variant is skipped too)."""
    aliases = {}
    for row in range(2, ws.max_row + 1):
        fdc_id = _cell_str(ws, row, 1)
        if fdc_id is None:
            continue
        action = _cell_str(ws, row, 4) or ""
        if action.startswith("הוסר"):
            continue
        alias = _cell_str(ws, row, 3)
        if alias:
            aliases[fdc_id] = alias
    return aliases


def _upsert_food_master(db, ws_catalog, aliases_by_fdc):
    created, updated = 0, 0
    new_fdc_ids = set()
    unmapped_categories = set()

    for row in range(2, ws_catalog.max_row + 1):
        name_he = _cell_str(ws_catalog, row, 1)
        if name_he is None:  # blank padding rows at the end of the sheet
            continue

        category_he = _cell_str(ws_catalog, row, 2)
        mapped_category = CATEGORY_MAP.get(category_he)
        if mapped_category is None:
            unmapped_categories.add(category_he)
            raise ValueError(f"No CATEGORY_MAP entry for {category_he!r} (row {row})")

        fdc_id = _cell_str(ws_catalog, row, 16)
        if fdc_id is None:
            raise ValueError(f"Missing מזהה FDC for {name_he!r} (row {row}) -- upsert key required")
        new_fdc_ids.add(fdc_id)

        alias = aliases_by_fdc.get(fdc_id)
        fields = dict(
            canonical_name_he=name_he,
            category=mapped_category,
            category_en=category_he,
            calories_per_100g=_cell_float(ws_catalog, row, 4),
            protein_per_100g=_cell_float(ws_catalog, row, 5),
            fat_per_100g=_cell_float(ws_catalog, row, 6),
            carbs_per_100g=_cell_float(ws_catalog, row, 7),
            fiber_per_100g=_cell_float(ws_catalog, row, 8),
            sugar_g=_cell_float(ws_catalog, row, 9),
            sodium_mg=_cell_float(ws_catalog, row, 10),
            potassium_mg=_cell_float(ws_catalog, row, 11),
            calcium_mg=_cell_float(ws_catalog, row, 12),
            iron_mg=_cell_float(ws_catalog, row, 13),
            saturated_fat_g=_cell_float(ws_catalog, row, 14),
            cholesterol_mg=_cell_float(ws_catalog, row, 15),
            source_url=_cell_str(ws_catalog, row, 18),
            barcode=_cell_str(ws_catalog, row, 19),
            # macros_estimated=False for every row in this file -- see the
            # module docstring: nothing here is FitAI's own calculation,
            # only USDA-sourced values (original, direct-field recovery, or
            # matched to an equivalent USDA SR Legacy/FNDDS record).
            macros_estimated=False,
            aliases=[alias] if alias and alias != name_he else [],
            is_active=True,
        )

        existing = db.query(FoodMaster).filter(FoodMaster.fdc_id == fdc_id).first()
        if existing:
            for key, value in fields.items():
                setattr(existing, key, value)
            updated += 1
        else:
            db.add(FoodMaster(fdc_id=fdc_id, created_by_user_id=None, **fields))
            created += 1

    return created, updated, new_fdc_ids


def _deactivate_not_in_new_set(db, new_fdc_ids):
    """Every existing active row not in this file's fdc_id set gets
    soft-deactivated, except the 9 PRESERVED_HE_NAMES legacy rows (fdc_id
    IS NULL, matched by canonical_name_he) -- same carve-out as
    seed_food_master_v2.py, confirmed 2026-08-01 for this file too since
    none of the 9 have an equivalent anywhere in the new catalog."""
    deactivated, preserved = 0, 0
    for food in db.query(FoodMaster).filter(FoodMaster.is_active.is_(True)).all():
        if food.fdc_id is None and food.canonical_name_he in PRESERVED_HE_NAMES:
            preserved += 1
            continue
        if food.fdc_id not in new_fdc_ids:
            food.is_active = False
            deactivated += 1
    return deactivated, preserved


def _reload_food_portions(db, ws_portions):
    rows_by_fdc_id = {}
    unmapped_units = set()
    for row in range(2, ws_portions.max_row + 1):
        fdc_id = _cell_str(ws_portions, row, 1)
        if fdc_id is None:
            continue
        rows_by_fdc_id.setdefault(fdc_id, []).append(row)

    portions_created = 0
    for fdc_id, rows in rows_by_fdc_id.items():
        db.query(FoodPortion).filter(FoodPortion.fdc_id == fdc_id).delete()
        for row in rows:
            unit_he = _cell_str(ws_portions, row, 4)
            unit_en = UNIT_HE_TO_EN.get(unit_he, unit_he)
            if unit_he and unit_he not in UNIT_HE_TO_EN:
                unmapped_units.add(unit_he)
            db.add(
                FoodPortion(
                    fdc_id=fdc_id,
                    quantity=_cell_float(ws_portions, row, 3),
                    unit_he=unit_he,
                    unit_en=unit_en or "unit",
                    description=_cell_str(ws_portions, row, 5),
                    weight_grams=_cell_float(ws_portions, row, 6),
                )
            )
            portions_created += 1

    return portions_created, len(rows_by_fdc_id), unmapped_units


def seed(xlsx_path: str, dry_run: bool = False, db=None) -> None:
    """db is injectable for tests/dry-runs against a throwaway engine; the
    CLI entrypoint always uses the real SessionLocal and owns its
    lifecycle."""
    owns_session = db is None
    if db is None:
        db = SessionLocal()
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws_catalog = wb[SHEET_CATALOG]
        ws_portions = wb[SHEET_PORTIONS]
        ws_aliases = wb[SHEET_ALIASES]

        aliases_by_fdc = _load_aliases(ws_aliases)

        created, updated, new_fdc_ids = _upsert_food_master(db, ws_catalog, aliases_by_fdc)
        deactivated, preserved = _deactivate_not_in_new_set(db, new_fdc_ids)
        portions_created, portion_fdc_ids, unmapped_units = _reload_food_portions(db, ws_portions)

        if dry_run:
            db.rollback()
        else:
            db.commit()

        print(f"food_master: created={created} updated={updated} deactivated={deactivated} preserved_legacy={preserved}")
        print(f"catalog fdc_ids: {len(new_fdc_ids)} (expected 334)")
        print(f"final active total: {len(new_fdc_ids) + preserved} (expected 343)")
        print("macros_estimated=true: 0 (hardcoded false -- see module docstring)")
        print(f"food_portions: created={portions_created} rows across {portion_fdc_ids} fdc_id(s)")
        print(f"aliases populated: {sum(1 for fdc in new_fdc_ids if fdc in aliases_by_fdc)}")
        if unmapped_units:
            print(f"UNIT_HE_TO_EN fallback (Hebrew reused as unit_en) for: {sorted(unmapped_units)}")
        if dry_run:
            print("\n[dry-run] no changes were committed.")
    finally:
        if owns_session:
            db.close()


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    dry_run = "--dry-run" in sys.argv
    if len(args) != 1:
        print("Usage: python seed_food_master_israel_v3.py <xlsx_path> [--dry-run]")
        sys.exit(1)
    seed(args[0], dry_run=dry_run)
